"""
generate_excel_report.py
------------------------
Generates a formatted, professional Excel operational report using openpyxl.

Produces: excel_reports/Healthcare_Ops_Weekly_Report.xlsx
  Sheet 1 - Weekly KPIs        : styled table with conditional formatting
  Sheet 2 - Specialty Summary  : revenue and no-show breakdown with data bars
  Sheet 3 - Staff Performance  : SLA league table with colour-coded ratings
  Sheet 4 - Follow-Up Actions  : patients overdue for follow-up

Author : Jomi Jose
Project: Healthcare Operations Reporting Automation
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import os

os.makedirs("excel_reports", exist_ok=True)

# ── Colours ───────────────────────────────────────────────────────────────────
BLUE_DARK  = "1F5C8B"
BLUE_MID   = "2E86C1"
BLUE_LIGHT = "D6EAF8"
TEAL       = "0D7377"
GREEN      = "1E8449"
GREEN_L    = "D5F5E3"
AMBER      = "D4AC0D"
AMBER_L    = "FEF9E7"
RED        = "C0392B"
RED_L      = "FDEDEC"
GRAY_L     = "F2F3F4"
WHITE      = "FFFFFF"


def header_style(ws, row, col, text, bg=BLUE_DARK, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def data_style(ws, row, col, value, number_format=None, bg=None, bold=False, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(size=10, name="Calibri", bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def add_title_block(ws, title, subtitle):
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = title
    t.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    t.fill      = PatternFill("solid", fgColor=BLUE_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = subtitle
    s.font      = Font(size=10, color=BLUE_DARK, italic=True, name="Calibri")
    s.fill      = PatternFill("solid", fgColor=BLUE_LIGHT)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


# ── Sheet 1: Weekly KPIs ──────────────────────────────────────────────────────
def sheet_weekly_kpis(wb):
    df = pd.read_csv("data/cleaned/weekly_summary.csv").tail(12)
    ws = wb.create_sheet("Weekly KPIs")
    ws.sheet_view.showGridLines = False

    add_title_block(ws, "Weekly Operational KPIs", "Last 12 Weeks  |  Clinic Healthcare Operations Report")

    headers = ["Week", "Total Appts", "Completed", "No-Shows",
               "No-Show Rate %", "Completion %", "Revenue (AED)", "Avg Lead Days"]
    cols    = ["week", "total_appts", "completed", "no_shows",
               "no_show_rate", "completion_rate", "revenue_aed", "avg_lead_days"]

    for ci, h in enumerate(headers, 1):
        c = header_style(ws, 4, ci, h)
        c.border = thin_border()

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = GRAY_L if ri % 2 == 0 else WHITE
        for ci, col in enumerate(cols, 1):
            val = row[col]
            fmt = None
            if col == "revenue_aed":       fmt = '#,##0'
            elif col in ["no_show_rate","completion_rate"]: fmt = '0.0"%"'
            c = data_style(ws, ri, ci, val, number_format=fmt, bg=bg)
            c.border = thin_border()

    # Conditional formatting on no-show rate column (E)
    last_row = 4 + len(df)
    ws.conditional_formatting.add(
        f"E5:E{last_row}",
        ColorScaleRule(start_type="min", start_color=GREEN,
                       mid_type="percentile", mid_value=50, mid_color=AMBER,
                       end_type="max", end_color=RED)
    )

    col_widths = [20, 14, 14, 14, 16, 14, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Bar chart — weekly appointment volume
    chart_row = last_row + 3
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Weekly Appointment Volume"
    chart.y_axis.title = "Appointments"
    chart.x_axis.title = "Week"
    chart.width   = 20
    chart.height  = 12
    chart.style   = 10

    data = Reference(ws, min_col=2, min_row=4, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=5, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{chart_row}")

    print("  Sheet 1: Weekly KPIs")


# ── Sheet 2: Specialty Summary ────────────────────────────────────────────────
def sheet_specialty(wb):
    df = pd.read_csv("data/cleaned/specialty_summary.csv")
    ws = wb.create_sheet("Specialty Summary")
    ws.sheet_view.showGridLines = False

    add_title_block(ws, "Performance by Specialty", "Appointments · Revenue · No-Show Analysis")

    headers = ["Specialty", "Total Appts", "Completed", "No-Shows",
               "No-Show Rate %", "Total Revenue (AED)", "Avg Fee (AED)"]
    cols    = ["specialty","total_appts","completed","no_shows",
               "no_show_rate","total_revenue","avg_fee"]

    for ci, h in enumerate(headers, 1):
        c = header_style(ws, 4, ci, h, bg=TEAL)
        c.border = thin_border()

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = GRAY_L if ri % 2 == 0 else WHITE
        for ci, col in enumerate(cols, 1):
            val = row[col]
            fmt = None
            if "revenue" in col or "fee" in col: fmt = '#,##0'
            elif "rate" in col:                   fmt = '0.0"%"'
            c = data_style(ws, ri, ci, val, number_format=fmt, bg=bg)
            c.border = thin_border()

    last_row = 4 + len(df)
    # Data bar on revenue column
    ws.conditional_formatting.add(
        f"F5:F{last_row}",
        DataBarRule(start_type="min", start_value=0,
                    end_type="max", end_value=None,
                    color=BLUE_MID)
    )

    col_widths = [22, 14, 14, 14, 16, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print("  Sheet 2: Specialty Summary")


# ── Sheet 3: Staff Performance ────────────────────────────────────────────────
def sheet_staff(wb):
    df = pd.read_csv("data/cleaned/staff_summary.csv")
    ws = wb.create_sheet("Staff Performance")
    ws.sheet_view.showGridLines = False

    add_title_block(ws, "Staff Performance League Table", "SLA · Escalations · Follow-Up Completion")

    headers = ["Staff ID", "Name", "Role", "Appts Handled",
               "SLA Met %", "Escalations", "Follow-Up %", "Avg Handle (min)"]
    cols    = ["staff_id","staff_name","role","total_appts_handled",
               "avg_sla_pct","total_escalations","avg_followup_pct","avg_handle_min"]

    for ci, h in enumerate(headers, 1):
        c = header_style(ws, 4, ci, h, bg="4A235A")
        c.border = thin_border()

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = GRAY_L if ri % 2 == 0 else WHITE
        for ci, col in enumerate(cols, 1):
            val = row[col]
            fmt = '0.0"%"' if "pct" in col else None
            c = data_style(ws, ri, ci, val, number_format=fmt, bg=bg)
            c.border = thin_border()

    last_row = 4 + len(df)
    # Green/Amber/Red on SLA column
    ws.conditional_formatting.add(
        f"E5:E{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["90"],
                   fill=PatternFill("solid", fgColor=GREEN_L))
    )
    ws.conditional_formatting.add(
        f"E5:E{last_row}",
        CellIsRule(operator="between", formula=["80","89.9"],
                   fill=PatternFill("solid", fgColor=AMBER_L))
    )
    ws.conditional_formatting.add(
        f"E5:E{last_row}",
        CellIsRule(operator="lessThan", formula=["80"],
                   fill=PatternFill("solid", fgColor=RED_L))
    )

    col_widths = [12, 14, 22, 16, 12, 14, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print("  Sheet 3: Staff Performance")


# ── Sheet 4: Follow-Up Actions ────────────────────────────────────────────────
def sheet_followup(wb):
    appts = pd.read_csv("data/cleaned/appointments_clean.csv",
                        parse_dates=["appointment_date"])
    pats  = pd.read_csv("data/cleaned/patients_clean.csv")

    overdue = appts[
        (appts["follow_up_needed"]=="Yes") &
        (appts["status"]=="Completed")
    ].merge(pats[["patient_id","first_name","last_name","preferred_channel"]],
            on="patient_id", how="left")

    overdue["days_since"] = (pd.Timestamp("2024-12-31") - overdue["appointment_date"]).dt.days
    overdue = overdue.sort_values("days_since", ascending=False).head(100)

    ws = wb.create_sheet("Follow-Up Actions")
    ws.sheet_view.showGridLines = False

    add_title_block(ws, "Follow-Up Action List", "Patients Due for Follow-Up — Sorted by Days Overdue")

    headers = ["Appt ID","Patient ID","Patient Name","Specialty",
               "Appointment Date","Days Since Visit","Preferred Channel","Action"]
    for ci, h in enumerate(headers, 1):
        c = header_style(ws, 4, ci, h, bg=RED)
        c.border = thin_border()

    for ri, (_, row) in enumerate(overdue.iterrows(), 5):
        bg = GRAY_L if ri % 2 == 0 else WHITE
        vals = [
            row["appointment_id"], row["patient_id"],
            f"{row['first_name']} {row['last_name']}",
            row["specialty"], str(row["appointment_date"].date()),
            int(row["days_since"]), row["preferred_channel"], "Call / Message"
        ]
        for ci, val in enumerate(vals, 1):
            c = data_style(ws, ri, ci, val, bg=bg)
            c.border = thin_border()

    # Highlight urgent (>60 days) in red
    last_row = 4 + len(overdue)
    ws.conditional_formatting.add(
        f"F5:F{last_row}",
        CellIsRule(operator="greaterThan", formula=["60"],
                   fill=PatternFill("solid", fgColor=RED_L))
    )

    col_widths = [14, 12, 22, 22, 18, 16, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print("  Sheet 4: Follow-Up Actions")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n── Generating Excel operational report ──────────────────────")
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    sheet_weekly_kpis(wb)
    sheet_specialty(wb)
    sheet_staff(wb)
    sheet_followup(wb)

    path = "excel_reports/Healthcare_Ops_Weekly_Report.xlsx"
    wb.save(path)
    print(f"\n  ✓ Saved: {path}\n")
